''' Mean Reward, Entropy and Loss from RLLIB's output

The calculation of the mean and confidence interval is done implicitly by the Seaborn lineplot function.
For each lineplot call, Seaborn takes the input data and calculates the mean and confidence interval automatically. The ci="sd" parameter tells Seaborn to use the standard deviation as the measure of uncertainty (confidence interval) for the mean.
The mean is calculated for each unique value of the x-axis (in this case, training_iteration) and is plotted as the line. The confidence interval is displayed as a shaded area around the mean line.
'''

import datetime
import os
import pandas as pd
import seaborn as sns
import matplotlib
matplotlib.use('Agg')  # Use the Agg backend
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from Z_utils import get_latest_file

current_dir = os.path.dirname(os.path.realpath(__file__))
TIMESTAMP = datetime.datetime.now().strftime("%Y%m%d-%H%M")

def graph_reward_n_others(rew_good_action = 10,rew_bad_action = -100 ):
    '''Plot mean and variance across several seeds'''

    excel_path = get_latest_file(directory = os.path.join(current_dir, 'A_results'), prefix = 'output', extension = '.xlsx') # Get the latest file in the directory by time stamp

    # Delete useless "Sheet1"
    book = load_workbook(excel_path)
    sheets = book.sheetnames
    if 'Sheet' in sheets:
        # If yes, remove "Sheet1"
        std = book['Sheet']
        book.remove(std)
    book.save(excel_path)

    # Read all sheets from the Excel file
    sheets = pd.read_excel(excel_path, sheet_name=None)

    # Create dictionaries to hold dataframes for each metric of each policy
    policy_reward_dfs,policy_loss_dfs,policy_entropy_dfs= {}, {}, {}
    unique_policies = set()  # Initialize a set to store unique policy names

    for sheet_name, df in sheets.items():
        df = df.set_index('training_iteration')

        # Dynamically identify policy columns and process data
        for col in df.columns:
            if 'custom_metrics/' in col:
                policy_name = col.split('/')[1]
                unique_policies.add(policy_name)  # Add policy name to the set
                temp_df = df[[col]].rename(columns={col: 'Reward'}).reset_index()
                temp_df['Policy'] = policy_name
                if policy_name in policy_reward_dfs:
                    policy_reward_dfs[policy_name] = pd.concat([policy_reward_dfs[policy_name], temp_df])
                else:
                    policy_reward_dfs[policy_name] = temp_df
            elif 'learner_stats/total_loss' in col:
                policy_name = col.split('/')[2]
                temp_df = df[[col]].rename(columns={col: 'Loss'}).reset_index()
                temp_df['Policy'] = policy_name
                if policy_name in policy_loss_dfs:
                    policy_loss_dfs[policy_name] = pd.concat([policy_loss_dfs[policy_name], temp_df])
                else:
                    policy_loss_dfs[policy_name] = temp_df
            elif 'learner_stats/entropy' in col:
                policy_name = col.split('/')[2]
                temp_df = df[[col]].rename(columns={col: 'Entropy'}).reset_index()
                temp_df['Policy'] = policy_name
                if policy_name in policy_entropy_dfs:
                    policy_entropy_dfs[policy_name] = pd.concat([policy_entropy_dfs[policy_name], temp_df])
                else:
                    policy_entropy_dfs[policy_name] = temp_df


    # Calculate Max Theoretical Reward per agent: max theo rew = 160.
    # num agents = 5, nbr coalitions each: 2^(n-1) = 2^4 = 16, rew_good_action: = 10
    number_of_policies = len(unique_policies)  # Number of unique policies
    max_theoretical_reward = number_of_policies * (2 ** (number_of_policies - 1)) * rew_good_action
    max_theoretical_reward = (2 ** (number_of_policies - 1)) * rew_good_action

    # Calculate AVG Reward Random Policy (only takes 2 actions, so 50% chances each)
    prob = 0.5
    rnd_pol_rew = prob * rew_good_action + prob * rew_bad_action


    # Plotting and Saving the Plots ===============================

    # Define metrics and their corresponding dataframes
    metrics_data = {
        'Reward': policy_reward_dfs,
        'Loss': policy_loss_dfs,
        'Entropy': policy_entropy_dfs
    }
    sns.set_style("whitegrid")
    for metric, dfs in metrics_data.items():
        plt.ioff()  #Turn off interactive mode - do not display graphs
        fig, ax = plt.subplots()
        for policy, data in dfs.items():
            sns.lineplot(x="training_iteration", y=metric, data=data, errorbar='sd', label=policy, ax=ax)

        if metric == 'Reward':
            # Add Max Theoretical Reward line
            max_reward_data = pd.DataFrame({
                'training_iteration': dfs[next(iter(dfs))]['training_iteration'],
                'Max Theoretical Return': [max_theoretical_reward] * len(dfs[next(iter(dfs))])
            })
            sns.lineplot(x='training_iteration',
                         y='Max Theoretical Return',
                         data=max_reward_data,
                         ax=ax,
                         label='Max Theoretical Return',
                         linestyle='--',  linewidth=2.5,    # Increase the line width
                         color='red')
            # Add Random Policy AVG Reward line
            rnd_reward_data = pd.DataFrame({
                'training_iteration': dfs[next(iter(dfs))]['training_iteration'],
                'Random Policy AVG Return': [rnd_pol_rew] * len(dfs[next(iter(dfs))])
            })
            sns.lineplot(x='training_iteration',
                        y='Random Policy AVG Return',
                        data=rnd_reward_data,
                        ax=ax,
                        label='Random Policy Return',
                        linestyle='--',
                        linewidth=2.5,
                        color='hotpink')

        if metric =='Reward':  metric = 'Returns' # What RLLIB calls 'Rewards' are actually "Undiscounted Returns"
        ax.set_xlabel("Training Iteration")
        ax.set_ylabel(f"{metric}")
        ax.set_title(f" Training {metric} by Policy")
        plt.legend()
        fig.savefig(os.path.join(current_dir, f"A_results/{metric.lower()}_plot_{TIMESTAMP}.png"), bbox_inches='tight')
        plt.close(fig)


##################################
if __name__ == "__main__":
    graph_reward_n_others()
    print('done!')